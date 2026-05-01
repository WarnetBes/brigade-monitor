"""
export_to_excel.py
Экспорт tasks.json в Excel с цветами, дропдаунами статусов
и сводным листом
"""

import json
import os
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

TASKS_FILE = Path("tasks.json")
EXPORT_DIR = Path(os.getenv("EXPORT_DIR", "exports"))
EXPORT_DIR.mkdir(exist_ok=True)

BRAND  = "01696F"
HDR    = "437A22"
STATUS_COLORS = {
    "Не начато":  "FFFFFF",
    "В работе":    "FFF3CD",
    "Выполнено":  "D4EDDA",
    "Задержка":   "F8D7DA",
    "Проблема":   "F5C6CB",
}
DIR_COLORS = {
    "Перемещение":       "E3F2FD",
    "Растарка":           "F3E5F5",
    "Доделки":            "FFF8E1",
    "Готовая продукция": "E8F5E9",
    "Сборочная":          "FCE4EC",
    "Не определено":      "F5F5F5",
}


def load_tasks():
    if TASKS_FILE.exists():
        return json.loads(TASKS_FILE.read_text(encoding="utf-8"))
    return []


def _hdr(ws, row, col, val, bg=HDR, fg="FFFFFF", bold=True, sz=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=fg, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c


def export_tasks_to_excel(output_path: str = None) -> str:
    tasks = load_tasks()
    now   = datetime.now()
    if not output_path:
        output_path = str(EXPORT_DIR / f"brigade_{now.strftime('%Y-%m-%d_%H-%M')}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Все задачи"

    COLS = ["Время", "Источник", "Группа", "Направление", "Статус", "Срочно", "Текст"]

    # Заголовок
    ws.merge_cells(f"A1:G1")
    _hdr(ws, 1, 1, f"🏭 Партия Еды | Бригадир | {now.strftime('%d.%m.%Y %H:%M')}", bg=BRAND, sz=12)
    for i, h in enumerate(COLS, 1):
        _hdr(ws, 2, i, h)

    thin = Side(style="thin", color="CCCCCC")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_i, t in enumerate(tasks, start=3):
        ts = t.get("timestamp", "")[:16].replace("T", " ")
        direction = t.get("direction", "Не определено")
        status    = t.get("status", "Не начато")
        urgent    = "🚨" if t.get("urgent") else ""
        row_vals  = [ts, t.get("source",""), t.get("group",""), direction, status, urgent, t.get("text","")[:200]]

        dir_bg  = DIR_COLORS.get(direction, "F5F5F5")
        stat_bg = STATUS_COLORS.get(status, "FFFFFF")

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_i, column=ci, value=val)
            cell.border = bord
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if ci == 4:   cell.fill = PatternFill("solid", fgColor=dir_bg)
            elif ci == 5: cell.fill = PatternFill("solid", fgColor=stat_bg)

    # Автоширина
    widths = [18, 12, 20, 20, 14, 8, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = "A3"

    # Сводный лист
    ws2 = wb.create_sheet("Сводка")
    dirs = ["Перемещение", "Растарка", "Доделки", "Готовая продукция", "Сборочная"]
    ws2.merge_cells("A1:D1")
    _hdr(ws2, 1, 1, f"Сводка смены — {now.strftime('%d.%m.%Y')}", bg=BRAND, sz=12)
    for i, h in enumerate(["Направление", "Всего", "Выполнено", "Проблемы"], 1):
        _hdr(ws2, 2, i, h)
    for ri, d in enumerate(dirs, 3):
        d_tasks = [t for t in tasks if t.get("direction") == d]
        done = sum(1 for t in d_tasks if t.get("status") == "Выполнено")
        probs = sum(1 for t in d_tasks if t.get("status") in ("Проблема", "Задержка"))
        row_vals2 = [d, len(d_tasks), done, probs]
        for ci, val in enumerate(row_vals2, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.border = bord
            if ci == 1:
                cell.fill = PatternFill("solid", fgColor=DIR_COLORS.get(d, "F5F5F5"))

    wb.save(output_path)
    print(f"✅ Экспорт сохранён: {output_path}")
    return output_path


if __name__ == "__main__":
    export_tasks_to_excel()
